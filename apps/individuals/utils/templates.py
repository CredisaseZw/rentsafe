import csv
import os
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.utils import quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from apps.individuals.models.models import Individual
from apps.common.models.models import Address, Suburb

def download_csv_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="individual_template.csv"'    
    writer = csv.writer(response)

    writer.writerow(["First Name*", "Last Name*", "Identification Type*", 
                     "ID Number*", "Marital Status", "date Of Birth", 
                     "Phone Number*", "Email Address","Address Type", "House/Flat Number", 
                     "Building/Complex Name", "Street Number*", "Street Name*", 
                     "City/Town*", "Province", "Country", "Area Code", "Current Employer", 
                     "Job Title", "Date Of Employment"])
    writer.writerow([
        "John", "Doe", "National ID", "63-1234567-A-00",
        "Married", "1985-07-15", "+263712345678", "john.doe@example.com","physical",
        "12B", "Greenwood Heights", "102", "Kwame Nkrumah Ave", "Bulawayo",
        "Bulawayo", "Zimbabwe", "00263","Tech Solutions", "Software Engineer", "2010-05-20"
    ])
    return response

def download_excel_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Individuals"

    # Define headers
    headers = [
        "First Name*", "Last Name*", "Date of Birth (YYYY-MM-DD)",
        "Gender", "Identification Type*", "Identification Number*", 
        "Marital Status", "Phone Number*", "Email Address", "Address Type",
        "House/Flat Number", "Building/Complex Name", "Street Number*", 
        "Street Name*","Suburb", "City/Town*", "Province", "Country", "Postal Code", 
        "Current Employer", "Job Title", "Date Of Employment (YYYY-MM-DD)"
    ]
    ws.append(headers)

    gender_choices = [choice[0] for choice in Individual.GENDER_CHOICES]
    id_type_choices = [choice[0] for choice in Individual.IDENTIFICATION_TYPES]
    marital_choices = [choice[0] for choice in Individual.MARITAL_STATUS_CHOICES]
    address_type_choices = [choice[0] for choice in Address.ADDRESS_TYPES]

    
    hidden = wb.create_sheet("Dropdowns")
    for i, val in enumerate(gender_choices, start=1):
        hidden[f"A{i}"] = val
    for i, val in enumerate(id_type_choices, start=1):
        hidden[f"B{i}"] = val
    for i, val in enumerate(marital_choices, start=1):
        hidden[f"C{i}"] = val
    for i, val in enumerate(address_type_choices, start=1):
        hidden[f"D{i}"] = val

    lookup_sheet = wb.create_sheet(title='Address-Lookups')
    lookup_sheet.append(["Suburb", "City", "Province", "Country"])
    
    suburbs = Suburb.objects.select_related("city__province__country")
    for i, suburb in enumerate(suburbs, start=2):
        lookup_sheet.append([
            suburb.name,
            suburb.city.name,
            suburb.city.province.name,
            suburb.city.province.country.name,
        ])    
    print("Total suburbs found:", suburbs.count())
    # Named range for suburb list
    suburb_range = f"{quote_sheetname('Address-Lookups')}!$A$2:$A${len(suburbs)+1}"
    dn = DefinedName(name="SuburbList", attr_text=suburb_range)
    wb.defined_names.add(dn)
    
    gender_dv = DataValidation(type="list", formula1="=Dropdowns!$A$1:$A$3", allow_blank=False)
    id_dv = DataValidation(type="list", formula1="=Dropdowns!$B$1:$B$4", allow_blank=False)
    marital_dv = DataValidation(type="list", formula1="=Dropdowns!$C$1:$C$4", allow_blank=True)
    address_type_dv = DataValidation(type="list", formula1="=Dropdowns!$D1:$D$5", allow_blank=False)
    suburb_dv= DataValidation(type="list",formula1="=SuburbList", allow_blank=False)
    

    ws.add_data_validation(gender_dv)
    ws.add_data_validation(id_dv)
    ws.add_data_validation(marital_dv)
    ws.add_data_validation(address_type_dv)
    ws.add_data_validation(suburb_dv)

    gender_dv.add("D2:D100")
    id_dv.add("E2:E100")
    marital_dv.add("G2:G100")
    address_type_dv.add("K2:K100")
    suburb_dv.add("P2:P100")
    
    # Auto-fill lookup formulas for city, province, country

    for row in range(2, 101):
        ws[f"Q{row}"] = f'=IFERROR(VLOOKUP(P{row}, \'Address-Lookups\'!$A$2:$D${len(suburbs)+1}, 2, FALSE), "")'
        ws[f"R{row}"] = f'=IFERROR(VLOOKUP(P{row}, \'Address-Lookups\'!$A$2:$D${len(suburbs)+1}, 3, FALSE), "")'
        ws[f"S{row}"] = f'=IFERROR(VLOOKUP(P{row}, \'Address-Lookups\'!$A$2:$D${len(suburbs)+1}, 4, FALSE), "")'
        


    hidden.sheet_state = 'hidden'
    lookup_sheet.sheet_state='hidden'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="individual_template.xlsx"'
    wb.save(response)
    return response