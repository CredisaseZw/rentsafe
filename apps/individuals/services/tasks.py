# apps/companies/tasks.py
import re
import csv
import secrets
import logging
from datetime import datetime
from celery import shared_task
from django.conf import settings
from django.db import transaction
from openpyxl import load_workbook
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from apps.common.models.models import Address, Suburb
from apps.common.services.tasks import send_notification
from django.contrib.contenttypes.models import ContentType
from apps.individuals.services.validators import validate_email, validate_national_id, normalize_zimbabwe_mobile
from apps.individuals.models.models import Individual, IndividualContactDetail,NextOfKin,EmploymentDetail

logger = logging.getLogger('individuals')


@shared_task
def send_individual_notification(individual_id: int, notification_type: str, context: dict, sender_id: int = None):
    """
    Send notification to individual
    
    Args:
        individual_id: ID of the individual
        notification_type: Type of notification
        context: Template context
        sender_id: ID of the sender
    """
    send_notification.delay(
        recipient_type='individual',
        recipient_id=individual_id,
        notification_type=notification_type,
        context=context,
        sender_id=sender_id,
    )
   
@shared_task    
def create_individual_background(self,individual_data: dict, user_id:int, request_path:str=None):
    """
    Create Individual in background after validation
    
    Args:
        individual_data: Validated individual data from serializer
        user_id: ID of the user creating the individual
        request_path: Request path for generating links
    """
    try:
        
        with transaction.atomic():
            address_data = individual_data.pop('addresses', [])
            employment_data = individual_data.pop('employment_details', [])
            kin_data = individual_data.pop('next_of_kin', [])
            contact_data= individual_data.pop('contact_details', [])
            
            individual = Individual.objects.create(**individual_data)
            
            logger.info(f"Created Individual {individual.id} in the background")
            
           # Create addresses
            for address_data in address_data:
                Address.objects.create(
                    content_object=individual,
                    **address_data
                )
            for emp_data in employment_data:
                EmploymentDetail.objects.create(
                    content_object = individual,
                    **emp_data
                )
                
            for kin in kin_data:
                NextOfKin.objects.create(
                    content_object = individual,
                    **kin
                )
            for contact in contact_data:
                IndividualContactDetail.objects.create(
                    content_object= individual,
                    **contact
                )
            
            otp_code = str(secrets.randbelow(999999)).zfill(6)
            individual_email= contact_data.email if contact_data else None

            if individual_email:
                # Send notification
                send_notification.delay(
                    recipient_type='individual',
                    recipient_id=individual.id,
                    notification_type=settings.ADD_INDIVIDUAL,
                    context={
                        'individual_name': f"{individual.first_name} {individual.last_name}",
                        'ID_number': individual.identification_number,
                        'user_id': user_id,
                    },
                    sender_id=user_id,
                    template_name='individual_registration',
                    subject='Individual Registration Successful - Fincheck',
                    include_otp=True,
                    otp_code=otp_code,
                    request_path=request_path,
                )
            return  {
                'success': True,
                'individual_id': individual.id,
                'message': 'Individual created successfully'
            }
    except Exception as exc:
        logger.error(f"Individual creation failed: {exc}")
        if self.request.retries < self.max_retries:
            raise self.retry(countdown=60, exc=exc)
        return {
            'success': False,
            'error': str(exc)
        }

@shared_task
def process_individuals_csv(file_path):
    import os

    def parse_date(value):
        if not value or str(value).strip() == "":
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except ValueError:
                continue
        return None

    with open(file_path, newline='', encoding='utf-8') as csv_file:
        reader = csv.reader(csv_file)
        headers = next(reader)

        skipped_rows = []
        created_rows = []

        for row in reader:
            if not any(row):
                continue

            errors = []

            if len(row) < 22:
                skipped_rows.append(row + ["Row too short"])
                continue

            first_name = row[0].strip()
            last_name = row[1].strip()
            dob = parse_date(row[2].strip())
            gender = row[3].strip()
            id_type = row[4].strip().lower()
            id_number = row[5].strip().upper()
            marital_status = row[6].strip()
            phone = row[7].strip()
            email = row[8].strip()
            address_type = row[9].strip()
            house_number = row[10].strip()
            building = row[11].strip()
            street_number = row[12].strip()
            street_name = row[13].strip()
            suburb_name = row[14].strip()
            city = row[15].strip()
            province = row[16].strip()
            country = row[17].strip()
            postal_code = row[18].strip()
            employer = row[19].strip()
            job_title = row[20].strip()
            employment_date = parse_date(row[21].strip())

            # Required fields check
            if not first_name or not last_name or not dob or not id_number:
                errors.append("Missing required fields")

            # Duplicate check
            if Individual.objects.filter(identification_number=id_number).exists():
                errors.append("Individual with this Identification number already exists.")

            # Validate national id or passport
            valid_national_id = False
            national_id_list = ["nationalid", "national id", "national_id"]
            if id_type in national_id_list:
                id_type = "national_id"
                if not id_number:
                    errors.append("Missing identification number")
                else:
                    try:
                        validate_national_id(id_number, "zimbabwe")
                        valid_national_id = True
                    except (ValidationError, ValueError) as e:
                        errors.append(str(e))
            elif id_type == "passport":
                id_type = "passport"
                valid_national_id = True
            else:
                errors.append("Invalid identification type")

            #validate phone 
            normalized_phone = normalize_zimbabwe_mobile(phone)
            if normalized_phone:
                phone = normalized_phone
            else:
                errors.append("Invalid mobile number")

            # Validate email
            if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                email = email
            else:
                errors.append("Invalid Email address")

            if not errors and valid_national_id:
                try:
                    # Create Individual
                    individual = Individual.objects.create(
                        first_name=first_name,
                        last_name=last_name,
                        date_of_birth=dob,
                        gender=gender,
                        identification_type=id_type,
                        identification_number=id_number,
                        marital_status=marital_status,
                    )
                    # Create contact info
                    IndividualContactDetail.objects.create(
                        individual=individual,
                        mobile_phone=phone,
                        email=email
                    )
                    # Create employment info
                    EmploymentDetail.objects.create(
                        individual=individual,
                        employer_name=employer,
                        job_title=job_title,
                        start_date=employment_date,
                    )
                    # Get or create Suburb
                    suburb, _ = Suburb.objects.get_or_create(
                        name=suburb_name,
                        city__name=city,
                        city__province__name=province,
                        city__province__country__name=country
                    )
                    # Create Address
                    Address.objects.create(
                        content_object=individual,
                        object_id=individual.pk,
                        address_type=address_type,
                        street_address=house_number,
                        line_2=building,
                        # street_number=street_number,
                        # street_name=street_name,
                        suburb=suburb,
                        postal_code=postal_code,
                    )

                    created_rows += 1
                except Exception as e:
                    errors.append(str(e))
                    skipped_rows.append(row + [", ".join(errors)])
            else:
                skipped_rows.append(row + [", ".join(errors)])

        # Write errors to file if any
        import os

        errors_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "errors")
        os.makedirs(errors_dir, exist_ok=True)
        if skipped_rows:
            error_file_path = os.path.join(errors_dir, f"errors_{os.path.basename(file_path)}")
            with open(error_file_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow([
                    'First Name', 'Last Name', 'Date Of Birth', 'Gender', 'Identification Type', 'ID Number',
                    'Marital Status', 'Phone Number', 'Email Address', 'Address Type', 'House/Flat Number',
                    'Building/Complex Name', 'Street Number', 'Street Name', 'Suburb', 'City/Town', 'Province',
                    'Country', 'Postal Code', 'Current Employer', 'Job Title', 'Date Of Employment', 'Errors'
                ])
                for row in skipped_rows:
                    writer.writerow(row)
            return {
                "status": "completed_with_errors",
                "created": len(created_rows),
                "skipped": len(skipped_rows),
                "error_file": error_file_path
            }
        else:
            return {
                "status": "completed",
                "created": len(created_rows),
                "skipped": 0
            }
                        

@shared_task
def process_individuals_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):

        first_name = row[0]
        last_name = row[1]
        dob = row[2]
        gender = row[3]
        id_type = row[4]
        id_number = row[5]
        marital_status = row[6]
        phone = row[7]
        email = row[8]
        address_type = row[9]
        house_number = row[10]
        building = row[11]
        street_number = row[12]
        street_name = row[13]
        suburb_name = row[14]
        city = row[15]
        province = row[16]
        country = row[17]
        postal_code = row[18]
        employer = row[19]
        job_title = row[20]
        employment_date = row[20]

        suburb, _ = Suburb.objects.get_or_create(
            name=suburb_name,
            city__name=city,
            city__province__name=province,
            city__province__country__name=country
        )

        # Create Address
        Address.objects.create(
            address_type=address_type,
            house_number=house_number,
            building_name=building,
            street_number=street_number,
            street_name=street_name,
            suburb=suburb,
            postal_code=postal_code,
        )

        # Create Individual
        individual = Individual.objects.create(
            first_name=first_name,
            last_name=last_name,
            date_of_birth=dob,
            gender=gender,
            identification_type=id_type,
            identification_number=id_number,
            marital_status=marital_status,
        )

        # Create contact info
        IndividualContactDetail.objects.create(
            individual= individual,
            mobile_phone=phone,
            email= email
        )

        # Create employment info
        EmploymentDetail.objects.create(
            individual= individual,
            current_employer=employer,
            job_title=job_title,
            date_of_employment=employment_date,
        )
    